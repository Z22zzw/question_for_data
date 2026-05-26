from pathlib import Path
import sqlite3

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import SESSION_COOKIE_NAME, create_app


def make_client(tmp_path: Path) -> TestClient:
    db_path = tmp_path / "test.sqlite"
    app = create_app(db_path=db_path, admin_password="secret")
    return TestClient(app)


def make_two_clients(tmp_path: Path) -> tuple[TestClient, TestClient]:
    db_path = tmp_path / "test.sqlite"
    app = create_app(db_path=db_path, admin_password="secret")
    return TestClient(app), TestClient(app)


def pretest_payload(version: str = "python") -> dict:
    payload = {
        "consent": "I agree",
        "questionnaire_version": version,
        "grade_year": "Year 3",
        "major": "计算机科学与技术" if version == "c" else "计算机类",
        "programming_experience_years": "3-4",
        "python_familiarity": "4",
        "file_io_familiarity": "3",
        "ai_tool_use_frequency": "Often",
        "ai_code_review_experience": "Sometimes",
    }
    if version == "python":
        payload["numpy_familiarity"] = "3"
    return payload


def start_session(client: TestClient):
    response = client.post("/api/session/start", json={"agreement": "I agree"})
    assert response.status_code == 200
    assert SESSION_COOKIE_NAME in response.cookies
    assert "httponly" in response.headers["set-cookie"].lower()
    return response


def submit_pretest(client: TestClient, version: str = "python") -> dict:
    pretest_response = client.post("/api/pretest", json=pretest_payload(version))
    assert pretest_response.status_code == 200
    assert pretest_response.json()["next_stage"] == "notice"
    response = start_session(client)
    assert response.json()["next_stage"] == "task"
    return response.json()


def correct_answers(start: int, end: int) -> dict:
    key = {
        "Q1": "B",
        "Q2": "B",
        "Q3": "B",
        "Q4": "A",
        "Q5": "B",
        "Q6": "B",
        "Q7": "B",
        "Q8": "D",
        "Q9": "A",
        "Q10": "B",
        "Q11": "A",
        "Q12": "A",
        "Q13": "A",
        "Q14": "A",
        "Q15": "D",
        "Q16": "B",
        "Q17": "B",
        "Q18": "B",
        "Q19": "A",
        "Q20": "B",
        "Q21": "A",
        "Q22": "A",
        "Q23": "A",
        "Q24": "A",
        "Q25": "D",
        "Q26": "B",
        "Q27": "B",
        "Q28": "B",
        "Q29": "A",
        "Q30": "A",
    }
    return {qid: answer for qid, answer in key.items() if start <= int(qid[1:]) <= end}


def posttest_payload() -> dict:
    return {
        "post_supervisor_role": "D",
        "post_requirements_first": "E",
        "post_missing_conditions": "D",
        "post_code_logic_tracing": "D",
        "post_output_prediction": "D",
        "post_test_design": "E",
        "post_human_intervention": "D",
        "post_responsible_submission": "E",
    }


def complete_all_tasks(client: TestClient) -> None:
    for task_id, start in enumerate([1, 6, 11, 16, 21, 26], start=1):
        response = client.post(
            f"/api/task/{task_id}",
            json={"answers": correct_answers(start, start + 4), "supervision_answers": {}},
        )
        assert response.status_code == 200


def mark_valid_completion_times(db_path: Path, session_id: str) -> None:
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            UPDATE sessions
            SET created_at = '2026-01-01T00:00:00+00:00',
                completed_at = '2026-01-01T00:12:30+00:00'
            WHERE id = ?
            """,
            (session_id,),
        )
        for task_id in range(1, 7):
            conn.execute(
                """
                UPDATE task_starts
                SET started_at = ?
                WHERE session_id = ? AND task_id = ?
                """,
                (f"2026-01-01T00:{(task_id - 1) * 2:02d}:00+00:00", session_id, task_id),
            )
            conn.execute(
                """
                UPDATE task_responses
                SET submitted_at = ?
                WHERE session_id = ? AND task_id = ?
                """,
                (f"2026-01-01T00:{(task_id - 1) * 2 + 1:02d}:30+00:00", session_id, task_id),
            )


def mark_session_started_at(db_path: Path, session_id: str, started_at: str) -> None:
    with sqlite3.connect(db_path) as conn:
        conn.execute("UPDATE sessions SET created_at = ? WHERE id = ?", (started_at, session_id))


def admin_sessions(client: TestClient) -> list[dict]:
    response = client.get("/api/admin/sessions?password=secret&include_abandoned=true")
    assert response.status_code == 200
    return response.json()


def latest_session_id(client: TestClient) -> str:
    return admin_sessions(client)[0]["session_id"]


def test_pretest_assigns_hidden_balanced_groups(tmp_path: Path):
    client = make_client(tmp_path)

    for idx in range(4):
        data = submit_pretest(client)
        assert data["next_task"] == 1
        assert data["next_stage"] == "task"
        assert data["participant_id"] == idx + 1
        assert set(data.keys()) == {
            "participant_id",
            "next_task",
            "next_stage",
            "time_limit_seconds",
            "remaining_seconds",
        }

    sessions = client.get("/api/admin/sessions?password=secret").json()
    groups = [row["group"] for row in sessions]
    assert groups.count("A") == 2
    assert groups.count("B") == 2


def test_python_and_c_versions_use_isolated_databases(tmp_path: Path):
    db_path = tmp_path / "test.sqlite"
    app = create_app(db_path=db_path, admin_password="secret")
    python_client = TestClient(app)
    c_client = TestClient(app)

    python_data = submit_pretest(python_client, "python")
    c_data = submit_pretest(c_client, "c")

    assert python_data["participant_id"] == 1
    assert c_data["participant_id"] == 1

    python_task = python_client.get("/api/task/1").json()
    c_task = c_client.get("/api/task/1").json()

    assert "Python" not in python_task["title"]
    assert c_task["title"].startswith("C Task 1")
    assert "double calculate_total" in c_task["code"]

    python_sessions = python_client.get("/api/admin/sessions?password=secret&version=python").json()
    c_sessions = python_client.get("/api/admin/sessions?password=secret&version=c").json()

    assert len(python_sessions) == 1
    assert len(c_sessions) == 1
    assert python_sessions[0]["questionnaire_version"] == "python"
    assert c_sessions[0]["questionnaire_version"] == "c"

    python_detail = python_client.get(
        f"/api/admin/sessions/{python_sessions[0]['session_id']}?password=secret&version=python"
    ).json()
    c_detail = python_client.get(
        f"/api/admin/sessions/{c_sessions[0]['session_id']}?password=secret&version=c"
    ).json()

    assert python_detail["pretest"]["numpy_familiarity"] == "3"
    assert "numpy_familiarity" not in c_detail["pretest"]


def test_agent_version_uses_isolated_database_and_agent_cards(tmp_path: Path):
    db_path = tmp_path / "test.sqlite"
    app = create_app(db_path=db_path, admin_password="secret")
    python_client = TestClient(app)
    agent_client_a = TestClient(app)
    agent_client_b = TestClient(app)

    python_data = submit_pretest(python_client, "python")
    agent_a_data = submit_pretest(agent_client_a, "agent")
    agent_b_data = submit_pretest(agent_client_b, "agent")

    assert python_data["participant_id"] == 1
    assert agent_a_data["participant_id"] == 1
    assert agent_b_data["participant_id"] == 2

    agent_task_a = agent_client_a.get("/api/task/1").json()
    agent_task_b = agent_client_b.get("/api/task/1").json()

    assert agent_task_a["title"].startswith("Agent Task 1")
    assert agent_task_a["questionnaire_version"] == "agent"
    assert "Agent action log" in agent_task_a["code"]
    assert "EN: gcc compiles" in agent_task_a["code"]
    assert "ZH: gcc 表示编译" in agent_task_a["code"]
    assert "--ids and --qty are test arguments" in agent_task_a["code"]
    assert agent_task_a["supervision_card"] is None
    assert agent_task_b["title"].startswith("Agent Task 1")
    assert agent_task_b["supervision_card"] is not None
    assert agent_task_b["supervision_card"][0]["dimension"] == "Factuality"

    python_sessions = python_client.get("/api/admin/sessions?password=secret&version=python").json()
    agent_sessions = python_client.get("/api/admin/sessions?password=secret&version=agent").json()
    agent_detail = python_client.get(
        f"/api/admin/sessions/{agent_sessions[0]['session_id']}?password=secret&version=agent"
    ).json()

    assert len(python_sessions) == 1
    assert len(agent_sessions) == 2
    assert agent_sessions[0]["questionnaire_version"] == "agent"
    assert "numpy_familiarity" not in agent_detail["pretest"]


def test_admin_bulk_delete_applies_only_to_selected_version(tmp_path: Path):
    db_path = tmp_path / "test.sqlite"
    app = create_app(db_path=db_path, admin_password="secret")
    client = TestClient(app)

    submit_pretest(client, "python")
    python_session_id = latest_session_id(client)
    client.post("/api/session/reset")
    submit_pretest(client, "c")
    c_session_id = client.get("/api/admin/sessions?password=secret&include_abandoned=true&version=c").json()[0]["session_id"]

    response = client.post(
        "/api/admin/sessions/bulk-delete?password=secret&version=c",
        json={"session_ids": [c_session_id]},
    )

    assert response.status_code == 200
    assert response.json()["deleted"] == 1
    assert client.get("/api/admin/sessions?password=secret&include_abandoned=true&version=c").json()[0]["status"] == "abandoned"
    assert client.get("/api/admin/sessions?password=secret&include_abandoned=true&version=python").json()[0]["session_id"] == python_session_id


def test_pretest_precedes_research_notice_agreement(tmp_path: Path):
    client = make_client(tmp_path)

    response = client.post("/api/pretest", json=pretest_payload())

    assert response.status_code == 200
    assert response.json()["next_stage"] == "notice"
    current = client.get("/api/session/current").json()
    assert current["status"] == "notice"
    assert current["remaining_seconds"] == 40 * 60
    start = start_session(client).json()
    assert start["next_stage"] == "task"
    assert start["remaining_seconds"] <= 40 * 60


def test_group_b_gets_supervision_cards_only_for_first_two_tasks(tmp_path: Path):
    client_a, client_b = make_two_clients(tmp_path)
    submit_pretest(client_a)
    submit_pretest(client_b)

    task1_a = client_a.get("/api/task/1").json()
    task1_b = client_b.get("/api/task/1").json()

    assert task1_a["supervision_card"] is None
    assert task1_b["supervision_card"] is not None

    client_b.post(
        "/api/task/1",
        json={"answers": correct_answers(1, 5), "supervision_answers": {"T1_SC_problem_definition": "Yes"}},
    )
    task2 = client_b.get("/api/task/2").json()
    assert task2["supervision_card"] is not None

    client_b.post(
        "/api/task/2",
        json={"answers": correct_answers(6, 10), "supervision_answers": {"T2_SC_problem_definition": "Yes"}},
    )
    task3 = client_b.get("/api/task/3").json()
    assert task3["supervision_card"] is None


def test_completed_task_cannot_be_read_or_modified(tmp_path: Path):
    client = make_client(tmp_path)
    submit_pretest(client)

    response = client.post(
        "/api/task/1",
        json={"answers": correct_answers(1, 5), "supervision_answers": {}},
    )
    assert response.status_code == 200

    back_read = client.get("/api/task/1")
    assert back_read.status_code == 409

    rewrite = client.post(
        "/api/task/1",
        json={"answers": {"Q1": "A"}, "supervision_answers": {}},
    )
    assert rewrite.status_code == 409


def test_scoring_and_excel_export(tmp_path: Path):
    client = make_client(tmp_path)
    pretest = submit_pretest(client)

    complete_all_tasks(client)

    summary_before_posttest = client.get("/api/session/current").json()
    assert summary_before_posttest["status"] == "posttest"
    assert summary_before_posttest["next_stage"] == "posttest"

    posttest_schema = client.get("/api/posttest?lang=en").json()
    assert posttest_schema["title"] == "AI Supervision Competence Post-test"
    assert len(posttest_schema["questions"]) == 8
    assert posttest_schema["questions"][0]["id"] == "post_supervisor_role"
    assert posttest_schema["questions"][0]["options"][0] == {"value": "E", "label": "Strongly agree"}
    assert posttest_schema["questions"][0]["options"][-1] == {"value": "A", "label": "Strongly disagree"}

    posttest_response = client.post("/api/posttest", json=posttest_payload())
    assert posttest_response.status_code == 200

    summary = client.get("/api/session/current").json()
    assert summary["status"] == "complete"
    assert summary["scores"]["total_score"] == 30
    assert summary["scores"]["deliverability_score"] == 12
    assert summary["scores"]["reasoning_score"] == 12
    assert summary["scores"]["error_identification_score"] == 6
    assert summary["posttest_scores"]["post_agent_supervision_score"] == 4.38

    mark_valid_completion_times(tmp_path / "test.sqlite", latest_session_id(client))
    export_response = client.get("/api/admin/export?password=secret")
    assert export_response.status_code == 200
    export_path = tmp_path / "export.xlsx"
    export_path.write_bytes(export_response.content)
    workbook = load_workbook(export_path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    row = [cell.value for cell in sheet[2]]
    data = dict(zip(headers, row))
    assert data["participant_id"] == pretest["participant_id"]
    assert data["total_score"] == 30
    assert data["post_supervisor_role"] == "D"
    assert data["post_responsible_submission"] == "E"
    assert data["post_agent_supervision_score"] == 4.38
    assert "total_duration_hms" in headers
    assert "task1_duration_hms" in headers
    assert "start_time" not in headers
    assert "pretest_submit_time" not in headers
    assert "end_time" not in headers


def test_admin_export_uses_selected_version_and_c_omits_numpy(tmp_path: Path):
    db_path = tmp_path / "test.sqlite"
    app = create_app(db_path=db_path, admin_password="secret")
    client = TestClient(app)

    python_pretest = submit_pretest(client, "python")
    complete_all_tasks(client)
    assert client.post("/api/posttest", json=posttest_payload()).status_code == 200
    python_session_id = latest_session_id(client)
    mark_valid_completion_times(db_path, python_session_id)

    c_pretest = submit_pretest(client, "c")
    complete_all_tasks(client)
    assert client.post("/api/posttest", json=posttest_payload()).status_code == 200
    c_sessions = client.get("/api/admin/sessions?password=secret&version=c").json()
    c_session_id = c_sessions[0]["session_id"]
    mark_valid_completion_times(db_path.with_name("test_c.sqlite"), c_session_id)

    python_export = client.get("/api/admin/export?password=secret&version=python")
    c_export = client.get("/api/admin/export?password=secret&version=c")

    assert python_export.status_code == 200
    assert c_export.status_code == 200
    assert "questionnaire_python_normal_completed_export.xlsx" in python_export.headers["content-disposition"]
    assert "questionnaire_c_normal_completed_export.xlsx" in c_export.headers["content-disposition"]

    python_path = tmp_path / "python_export.xlsx"
    c_path = tmp_path / "c_export.xlsx"
    python_path.write_bytes(python_export.content)
    c_path.write_bytes(c_export.content)

    python_sheet = load_workbook(python_path).active
    c_sheet = load_workbook(c_path).active
    python_headers = [cell.value for cell in python_sheet[1]]
    c_headers = [cell.value for cell in c_sheet[1]]
    python_row = dict(zip(python_headers, [cell.value for cell in python_sheet[2]]))
    c_row = dict(zip(c_headers, [cell.value for cell in c_sheet[2]]))

    assert "numpy_familiarity" in python_headers
    assert "numpy_familiarity" not in c_headers
    assert python_row["participant_id"] == python_pretest["participant_id"]
    assert c_row["participant_id"] == c_pretest["participant_id"]
    assert python_row["questionnaire_version"] == "python"
    assert c_row["questionnaire_version"] == "c"


def test_admin_export_uses_agent_version_database(tmp_path: Path):
    db_path = tmp_path / "test.sqlite"
    app = create_app(db_path=db_path, admin_password="secret")
    client = TestClient(app)

    python_pretest = submit_pretest(client, "python")
    complete_all_tasks(client)
    assert client.post("/api/posttest", json=posttest_payload()).status_code == 200
    mark_valid_completion_times(db_path, latest_session_id(client))

    agent_pretest = submit_pretest(client, "agent")
    complete_all_tasks(client)
    assert client.post("/api/posttest", json=posttest_payload()).status_code == 200
    agent_session_id = client.get("/api/admin/sessions?password=secret&version=agent").json()[0]["session_id"]
    mark_valid_completion_times(db_path.with_name("test_agent.sqlite"), agent_session_id)

    agent_export = client.get("/api/admin/export?password=secret&version=agent")

    assert agent_export.status_code == 200
    assert "questionnaire_agent_normal_completed_export.xlsx" in agent_export.headers["content-disposition"]

    export_path = tmp_path / "agent_export.xlsx"
    export_path.write_bytes(agent_export.content)
    sheet = load_workbook(export_path).active
    headers = [cell.value for cell in sheet[1]]
    row = dict(zip(headers, [cell.value for cell in sheet[2]]))

    assert "numpy_familiarity" not in headers
    assert row["participant_id"] == agent_pretest["participant_id"]
    assert row["participant_id"] != python_pretest["participant_id"] or row["questionnaire_version"] == "agent"
    assert row["questionnaire_version"] == "agent"


def test_timeout_marks_session_blocks_submissions_and_reports_duration(tmp_path: Path):
    db_path = tmp_path / "test.sqlite"
    app = create_app(db_path=db_path, admin_password="secret")
    client = TestClient(app)

    submit_pretest(client)
    session_id = latest_session_id(client)
    mark_session_started_at(db_path, session_id, "2026-01-01T00:00:00+00:00")

    current = client.get("/api/session/current").json()
    assert current["status"] == "timeout"
    assert current["next_stage"] is None
    assert current["remaining_seconds"] == 0
    assert current["elapsed_seconds"] >= 40 * 60

    pretest_response = client.post("/api/pretest", json=pretest_payload())
    assert pretest_response.status_code == 410
    assert pretest_response.json()["detail"] == "Session timed out. Please restart the questionnaire."

    row = admin_sessions(client)[0]
    assert row["completion_bucket"] == "incomplete"
    assert row["incomplete_reason"] == "timeout"
    assert row["total_duration_hms"].startswith("00:40:")


def test_admin_export_only_includes_normal_completed_sessions(tmp_path: Path):
    db_path = tmp_path / "test.sqlite"
    app = create_app(db_path=db_path, admin_password="secret")
    client = TestClient(app)

    normal = submit_pretest(client)
    complete_all_tasks(client)
    assert client.post("/api/posttest", json=posttest_payload()).status_code == 200
    normal_session_id = latest_session_id(client)
    mark_valid_completion_times(db_path, normal_session_id)

    pending_posttest = submit_pretest(client)
    complete_all_tasks(client)

    abandoned = submit_pretest(client)
    abandoned_session_id = latest_session_id(client)
    assert client.delete(f"/api/admin/sessions/{abandoned_session_id}?password=secret").status_code == 200

    invalid_complete = submit_pretest(client)
    complete_all_tasks(client)
    assert client.post("/api/posttest", json=posttest_payload()).status_code == 200

    sessions = admin_sessions(client)
    by_participant = {row["participant_id"]: row for row in sessions}
    assert by_participant[normal["participant_id"]]["completion_bucket"] == "normal_completed"
    assert by_participant[pending_posttest["participant_id"]]["completion_bucket"] == "incomplete"
    assert by_participant[pending_posttest["participant_id"]]["incomplete_reason"] == "pending_posttest"
    assert by_participant[abandoned["participant_id"]]["completion_bucket"] == "incomplete"
    assert by_participant[abandoned["participant_id"]]["incomplete_reason"] == "abandoned"
    assert by_participant[invalid_complete["participant_id"]]["completion_bucket"] == "incomplete"
    assert by_participant[invalid_complete["participant_id"]]["incomplete_reason"] == "quality_failed"
    assert by_participant[normal["participant_id"]]["total_duration_hms"] == "00:12:30"

    export_response = client.get("/api/admin/export?password=secret")
    assert export_response.status_code == 200
    export_path = tmp_path / "normal_completed_export.xlsx"
    export_path.write_bytes(export_response.content)
    workbook = load_workbook(export_path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    rows = [dict(zip(headers, [cell.value for cell in row])) for row in sheet.iter_rows(min_row=2)]

    assert [row["participant_id"] for row in rows] == [normal["participant_id"]]


def test_task_can_be_localized_to_chinese(tmp_path: Path):
    client = make_client(tmp_path)
    submit_pretest(client)

    task = client.get("/api/task/1?lang=zh").json()

    assert task["title"].startswith("任务 1")
    assert "人工智能编程智能体(AI coding agent)" in task["questions"][0]["prompt"]
    assert "代码(code)" in task["questions"][0]["prompt"]
    q5_options = [option["text"] for option in task["questions"][4]["options"]]
    assert q5_options == [
        "VIP 折扣计算错误",
        "未知商品 ID 被当作价格 0 处理",
        "没有保留两位小数",
        "运费规则完全反了",
    ]


def test_c_task_can_be_localized_to_chinese(tmp_path: Path):
    client_a, client_b = make_two_clients(tmp_path)
    submit_pretest(client_a, "c")
    submit_pretest(client_b, "c")

    task = client_b.get("/api/task/1?lang=zh").json()

    assert task["title"].startswith("C 任务 1")
    assert "商品编号不存在" in task["requirements"][2]
    assert task["questions"][0]["prompt"] == "这份人工智能编程智能体(AI coding agent)生成的 C 语言代码(code)是否完全满足任务要求？"
    assert task["questions"][4]["options"][1]["text"] == "未知商品编号被按 0 元处理"
    assert task["supervision_card"][0]["dimension"] == "理解任务要求"
    assert "返回 -1" in task["supervision_card"][0]["prompt"]


def test_posttest_is_same_for_a_and_b_and_available_after_tasks(tmp_path: Path):
    clients = make_two_clients(tmp_path)
    for client in clients:
        submit_pretest(client)

    posttest_titles = []
    for client in clients:
        for task_id, start in enumerate([1, 6, 11, 16, 21, 26], start=1):
            response = client.post(
                f"/api/task/{task_id}",
                json={"answers": correct_answers(start, start + 4), "supervision_answers": {}},
            )
            assert response.status_code == 200
        schema = client.get("/api/posttest?lang=zh").json()
        posttest_titles.append(schema["title"])
        assert schema["sections"][0]["title"] == "人工智能(AI)监督意识"
        assert len(schema["questions"]) == 8

    assert posttest_titles[0] == posttest_titles[1]


def test_admin_can_close_questionnaire_version_for_frontend(tmp_path: Path):
    client = make_client(tmp_path)

    settings = client.get("/api/questionnaire-settings").json()
    assert {item["version"]: item["enabled"] for item in settings["versions"]}["agent"] is True

    update = client.put(
        "/api/admin/questionnaire-settings?password=secret",
        json={"enabled_versions": {"agent": False}},
    )
    assert update.status_code == 200
    assert {item["version"]: item["enabled"] for item in update.json()["versions"]}["agent"] is False

    response = client.post("/api/pretest", json=pretest_payload("agent"))
    assert response.status_code == 403
    assert response.json()["detail"] == "This questionnaire version is currently closed"


def test_session_current_and_reset_use_cookie_not_local_storage_ids(tmp_path: Path):
    client = make_client(tmp_path)

    assert client.get("/api/session/current").json()["status"] == "none"

    pretest_response = client.post("/api/pretest", json=pretest_payload())
    assert pretest_response.status_code == 200
    assert SESSION_COOKIE_NAME in pretest_response.cookies

    current_notice = client.get("/api/session/current").json()
    assert current_notice["status"] == "notice"
    assert current_notice["next_stage"] == "notice"
    assert current_notice["time_limit_seconds"] == 40 * 60
    assert current_notice["remaining_seconds"] == 40 * 60

    start_response = start_session(client)
    assert pretest_response.cookies[SESSION_COOKIE_NAME] == client.cookies.get(SESSION_COOKIE_NAME)

    current = client.get("/api/session/current").json()
    assert current["status"] == "in_progress"
    assert current["next_task"] == 1
    assert "group" not in current
    assert "session_id" not in current

    reset_response = client.post("/api/session/reset")
    assert reset_response.status_code == 200
    assert client.get("/api/session/current").json()["status"] == "none"
